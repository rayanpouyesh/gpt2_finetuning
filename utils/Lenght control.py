import os
import pandas as pd
from openpyxl import load_workbook

# تابع برای تقسیم داده‌ها بر اساس طول جملات و ذخیره در ادامه فایل اکسل موجود
def split_sentences_by_length(input_file, output_file_short, output_file_long, length_threshold=256):
    # خواندن فایل اکسل
    df = pd.read_excel(input_file)

    # اطمینان از وجود ستون مورد نظر
    if 'origin' not in df.columns:
        raise ValueError("The input file must contain a column named 'sentence'")

    # فیلتر کردن جملات کوتاه‌تر و بلندتر از مقدار تعیین‌شده
    short_sentences = df[df['origin'].str.len() <= length_threshold]
    long_sentences = df[df['origin'].str.len() > length_threshold]

    # ذخیره داده‌های جملات کوتاه‌تر در ادامه فایل اکسل موجود
    if os.path.exists(output_file_short):
        with pd.ExcelWriter(output_file_short, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            short_sentences.to_excel(writer, index=False, startrow=writer.sheets['Sheet1'].max_row, header=False)
    else:
        short_sentences.to_excel(output_file_short, index=False)

    # ذخیره داده‌های جملات بلندتر در ادامه فایل اکسل موجود
    if os.path.exists(output_file_long):
        with pd.ExcelWriter(output_file_long, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            long_sentences.to_excel(writer, index=False, startrow=writer.sheets['Sheet1'].max_row, header=False)
    else:
        long_sentences.to_excel(output_file_long, index=False)

# مسیر فایل ورودی و خروجی
input_file = "C:\\rayanpoyesh_project\\Ghaemi\\summary.xlsx"  # فایل ورودی
output_file_short = "dataset1.xlsx"  # فایل جملات کوتاه‌تر
output_file_long = "512_sentences.xlsx"  # فایل جملات بلندتر

# اجرای تابع با مقدار پیش‌فرض طول 256
split_sentences_by_length(input_file, output_file_short, output_file_long)
